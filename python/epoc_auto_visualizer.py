#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
EPOC系Excelを自動判定して可視化済みExcelを出力する統合ツール。

対応形式:
1. 経験症候・疾患ファイル
   - 各研修医が1シート
   - 「経験すべき症候」「経験すべき疾患」ブロックあり
2. 評価ファイル
   - 「研修医評価」「指導医評価」シートあり

使い方:
    python epoc_auto_visualizer.py input.xlsx
"""

from __future__ import annotations

import sys
import json
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------- common ----------
BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
BOLD = Font(bold=True)

def cell_text(v) -> str:
    return "" if v is None else str(v).strip()

def safe_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except Exception:
        return 0

def safe_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def autosize_columns(ws):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        length = 0
        for cell in col_cells:
            try:
                length = max(length, len("" if cell.value is None else str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), 30)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = BOLD
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_numeric_fill(ws):
    headers = [cell_text(c.value) for c in ws[1]]
    numeric_like = {"経験数","承認数","病歴要約提出","病歴要約確認","外科手術要約提出","外科手術要約確認",
                    "評価点","評価点_num","平均点","平均点_全体","平均点_A","平均点_B","平均点_C","最低点"}
    target_cols = [i for i,h in enumerate(headers, start=1) if h in numeric_like or h.startswith(("A-","B-","C-"))]
    for row in ws.iter_rows(min_row=2):
        for idx in target_cols:
            cell = row[idx-1]
            try:
                v = float(cell.value)
            except Exception:
                continue
            if v == 0:
                cell.fill = RED_FILL
            elif v < 2:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = GREEN_FILL


# ---------- detector ----------
def detect_workbook_type(input_path: Path) -> str:
    wb = load_workbook(input_path, data_only=True)
    sheetnames = wb.sheetnames

    if "研修医評価" in sheetnames or "指導医評価" in sheetnames:
        return "evaluation"

    for ws in wb.worksheets:
        texts = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), max_col=min(ws.max_column, 10), values_only=True):
            texts.extend([cell_text(v) for v in row if cell_text(v)])
            # Fallback inner check for evaluation columns
            if "研修医UMIN ID" in texts and "研修医氏名" in texts:
                return "evaluation"
            # Fallback inner check for case summary columns
            if "体重減少・るい痩" in texts and "研修医氏名" in texts and "ショック" in texts:
                # 簡易版の場合、値が数値（症例数）ならsymptom_disease_simplified、未/済ならcase_summaryとするべきですが、
                # ヘッダー自体は同じなので、ここでは 'case_summary_or_simplified' のような中間判定も可能。
                # ただしユーザーアップロード時は --type で来るためファイル種別でカバー可能。
                # もし--type指定なしの場合は、ここで区別が難しいので case_summary をベースに返すか、判別ロジックを追加します。
                # 簡易的に、データ行(row>1)に「未」「済」があればcase_summary、数値だけなら症例とみなすなどの手段がありますが、
                # まずは "case_summary" or "symptom_disease_simplified" 判定のため、10行のテキストをチェックします。
                if "未" in texts or "済" in texts:
                    return "case_summary"
                else:
                    return "symptom_disease_simplified"
                
        joined = " ".join(texts)
        if "研修医氏名" in joined and "経験すべき症候" in joined:
            return "symptom_disease"

    return "unknown"


# ---------- symptom / disease ----------
SECTION_LABELS = {"症候": "経験すべき症候", "疾患": "経験すべき疾患"}

# Excel列AF(32列目)以降を疾患とする（症候=1-31列、疾患=32列〜）
SYMPTOM_DISEASE_BOUNDARY_COL = 32  # AF列

HEADER_ALIASES = {
    "item": ["症候名", "疾患名"],
    "experience_count": ["経験数"],
    "approved_count": ["承認数"],
    "history_submitted": ["病歴要約の提出の有無"],
    "history_confirmed": ["病歴要約の確認の有無"],
    "surgery_submitted": ["外科手術要約の提出の有無"],
    "surgery_confirmed": ["外科手術要約の確認の有無"],
}

def find_row_by_label(ws, label: str) -> Optional[int]:
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if cell_text(ws.cell(r, c).value) == label:
                return r
    return None

def find_metadata(ws) -> Tuple[str, str]:
    resident_name = ""
    umin_id = ""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        values = [cell_text(v) for v in row]
        for i, v in enumerate(values):
            if v == "研修医氏名" and i + 1 < len(values):
                resident_name = values[i + 1]
            if v == "研修医UMIN ID" and i + 1 < len(values):
                umin_id = values[i + 1]
    return resident_name, umin_id

def normalize_header(name: str) -> Optional[str]:
    t = cell_text(name)
    for key, aliases in HEADER_ALIASES.items():
        if t in aliases:
            return key
    return None

def parse_section(ws, resident_name: str, umin_id: str, section_type: str) -> List[Dict]:
    label = SECTION_LABELS[section_type]
    start_row = find_row_by_label(ws, label)
    if not start_row:
        return []
    header_row = start_row + 1
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        norm = normalize_header(ws.cell(header_row, c).value)
        if norm:
            col_map[norm] = c
    if not all(k in col_map for k in ["item", "experience_count", "approved_count"]):
        return []

    next_section_row = find_row_by_label(ws, SECTION_LABELS["疾患"]) if section_type == "症候" else None

    records = []
    row = header_row + 1
    while row <= ws.max_row:
        if next_section_row and row >= next_section_row:
            break
        item_name = cell_text(ws.cell(row, col_map["item"]).value)
        if not item_name:
            row += 1
            continue
        if item_name in SECTION_LABELS.values():
            break
        records.append({
            "研修医氏名": resident_name,
            "研修医UMIN ID": umin_id,
            "区分": section_type,
            "項目名": item_name,
            "経験数": safe_int(ws.cell(row, col_map["experience_count"]).value),
            "承認数": safe_int(ws.cell(row, col_map["approved_count"]).value),
            "病歴要約提出": safe_int(ws.cell(row, col_map.get("history_submitted", 0)).value if col_map.get("history_submitted") else 0),
            "病歴要約確認": safe_int(ws.cell(row, col_map.get("history_confirmed", 0)).value if col_map.get("history_confirmed") else 0),
            "外科手術要約提出": safe_int(ws.cell(row, col_map.get("surgery_submitted", 0)).value if col_map.get("surgery_submitted") else 0),
            "外科手術要約確認": safe_int(ws.cell(row, col_map.get("surgery_confirmed", 0)).value if col_map.get("surgery_confirmed") else 0),
            "シート名": ws.title,
        })
        row += 1
    return records

def parse_symptom_disease_workbook(input_path: Path) -> pd.DataFrame:
    wb = load_workbook(input_path, data_only=True)
    records = []
    for ws in wb.worksheets:
        resident_name, umin_id = find_metadata(ws)
        if not resident_name and not umin_id:
            continue
        records.extend(parse_section(ws, resident_name, umin_id, "症候"))
        records.extend(parse_section(ws, resident_name, umin_id, "疾患"))
    if not records:
        # Return an empty dataframe with expected columns if no data found instead of crashing
        return pd.DataFrame(columns=[
            "研修医氏名", "研修医UMIN ID", "区分", "項目名", "経験数", "承認数",
            "病歴要約提出", "病歴要約確認", "外科手術要約提出", "外科手術要約確認", "シート名"
        ])
    df = pd.DataFrame(records)
    ordered_cols = [
        "研修医氏名", "研修医UMIN ID", "区分", "項目名", "経験数", "承認数",
        "病歴要約提出", "病歴要約確認", "外科手術要約提出", "外科手術要約確認", "シート名"
    ]
    return df[ordered_cols]

def make_sd_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for resident, sub in df.groupby("研修医氏名", dropna=False):
        symptom = sub[sub["区分"] == "症候"]
        disease = sub[sub["区分"] == "疾患"]
        def rate(x):
            return round((x["経験数"] > 0).sum() / len(x), 3) if len(x) else 0.0
        total_items = len(sub)
        experienced_items = (sub["経験数"] > 0).sum()
        approved_items = (sub["承認数"] > 0).sum()
        alert_count = (
            ((sub["経験数"] == 0)).sum()
            + (((sub["経験数"] > 0) & (sub["承認数"] == 0))).sum()
            + (((sub["経験数"] > 0) & (sub["病歴要約提出"] == 0))).sum()
            + (((sub["病歴要約提出"] > 0) & (sub["病歴要約確認"] == 0))).sum()
        )
        rows.append({
            "研修医氏名": resident,
            "研修医UMIN ID": sub["研修医UMIN ID"].dropna().iloc[0] if not sub["研修医UMIN ID"].dropna().empty else "",
            "総項目数": total_items,
            "経験済項目数": experienced_items,
            "承認済項目数": approved_items,
            "症候達成率": rate(symptom),
            "疾患達成率": rate(disease),
            "全体達成率": round(experienced_items / total_items, 3) if total_items else 0.0,
            "承認率": round(approved_items / total_items, 3) if total_items else 0.0,
            "病歴要約未提出件数": (((sub["経験数"] > 0) & (sub["病歴要約提出"] == 0))).sum(),
            "病歴要約未確認件数": (((sub["病歴要約提出"] > 0) & (sub["病歴要約確認"] == 0))).sum(),
            "外科手術要約未提出件数": (((sub["経験数"] > 0) & (sub["外科手術要約提出"] == 0))).sum(),
            "外科手術要約未確認件数": (((sub["外科手術要約提出"] > 0) & (sub["外科手術要約確認"] == 0))).sum(),
            "要対応件数": int(alert_count),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["要対応件数", "全体達成率"], ascending=[False, True]).reset_index(drop=True)
    return out

def make_sd_alerts(df: pd.DataFrame) -> pd.DataFrame:
    alerts = []
    for _, r in df.iterrows():
        if r["経験数"] == 0:
            alerts.append({**r.to_dict(), "状態": "未経験"})
        elif r["経験数"] > 0 and r["承認数"] == 0:
            alerts.append({**r.to_dict(), "状態": "経験済だが承認待ち"})
        if r["経験数"] > 0 and r["病歴要約提出"] == 0:
            alerts.append({**r.to_dict(), "状態": "病歴要約未提出"})
        if r["病歴要約提出"] > 0 and r["病歴要約確認"] == 0:
            alerts.append({**r.to_dict(), "状態": "病歴要約未確認"})
        if r["経験数"] > 0 and r["外科手術要約提出"] == 0:
            alerts.append({**r.to_dict(), "状態": "外科手術要約未提出"})
        if r["外科手術要約提出"] > 0 and r["外科手術要約確認"] == 0:
            alerts.append({**r.to_dict(), "状態": "外科手術要約未確認"})
    out = pd.DataFrame(alerts)
    if not out.empty:
        keep = ["研修医氏名","研修医UMIN ID","区分","項目名","経験数","承認数","病歴要約提出","病歴要約確認","外科手術要約提出","外科手術要約確認","状態"]
        out = out[keep].sort_values(["研修医氏名","区分","項目名","状態"])
    return out

def df_to_records(df: pd.DataFrame) -> List[Dict]:
    if df.empty:
        return []
    return json.loads(df.to_json(orient="records", force_ascii=False))

def make_sd_items(raw_df: pd.DataFrame, is_simplified: bool = False) -> pd.DataFrame:
    """項目ごとの達成状況を算出。ダッシュボードで達成/未達成を一目で表示するため。"""
    if raw_df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in raw_df.iterrows():
        exp = safe_int(r.get("経験数", 0))
        appr = safe_int(r.get("承認数", 0))
        med_submit = safe_int(r.get("病歴要約提出", 0))
        med_confirm = safe_int(r.get("病歴要約確認", 0))
        surg_submit = safe_int(r.get("外科手術要約提出", 0))
        surg_confirm = safe_int(r.get("外科手術要約確認", 0))
        if exp == 0:
            status = "未経験"
        elif appr == 0:
            status = "承認待ち"
        elif is_simplified:
            status = "達成"
        elif exp > 0 and med_submit == 0:
            status = "病歴要約未提出"
        elif med_submit > 0 and med_confirm == 0:
            status = "病歴要約未確認"
        elif exp > 0 and surg_submit == 0:
            status = "外科手術要約未提出"
        elif surg_submit > 0 and surg_confirm == 0:
            status = "外科手術要約未確認"
        else:
            status = "達成"
        rows.append({
            "研修医氏名": r.get("研修医氏名"),
            "研修医UMIN ID": r.get("研修医UMIN ID"),
            "区分": r.get("区分"),
            "項目名": r.get("項目名"),
            "経験数": exp,
            "承認数": appr,
            "状態": status,
        })
    out = pd.DataFrame(rows)
    return out.sort_values(["研修医氏名", "区分", "項目名"]).reset_index(drop=True)

def pivot_sd_heatmap(df: pd.DataFrame, section: str) -> pd.DataFrame:
    sub = df[df["区分"] == section].copy()
    if sub.empty:
        return pd.DataFrame()
    pt = pd.pivot_table(sub, index="研修医氏名", columns="項目名", values="経験数", aggfunc="sum", fill_value=0)
    return pt.reset_index()

def write_sd_output(input_path: Path, raw_df, summary_df, alert_df, heat_symptom, heat_disease, items_df=None, is_simplified: bool = False) -> Tuple[Path, Path]:
    output_path = input_path.with_name(input_path.stem + "_visualized.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, index=False, sheet_name="raw_all")
        summary_df.to_excel(writer, index=False, sheet_name="summary_resident")
        alert_df.to_excel(writer, index=False, sheet_name="alert_list")
        heat_symptom.to_excel(writer, index=False, sheet_name="heatmap_symptom")
        heat_disease.to_excel(writer, index=False, sheet_name="heatmap_disease")
    wb = load_workbook(output_path)
    for name in ["raw_all","summary_resident","alert_list","heatmap_symptom","heatmap_disease"]:
        ws = wb[name]
        style_header(ws, 1)
        autosize_columns(ws)
        apply_numeric_fill(ws)
    ws = wb.create_sheet("dashboard")
    ws["A1"] = "EPOC 研修進捗ダッシュボード"; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "全体サマリー"; ws["A3"].font = BOLD
    stats = [
        ("研修医数", len(summary_df)),
        ("平均全体達成率", round(summary_df["全体達成率"].mean(), 3) if not summary_df.empty else 0),
        ("平承認率(簡易版は0)", round(summary_df["承認率"].mean(), 3) if "承認率" in summary_df.columns and not summary_df.empty else 0),
        ("要対応総件数", len(alert_df)),
    ]
    r = 4
    for k,v in stats:
        ws[f"A{r}"] = k; ws[f"B{r}"] = v; r += 1
    autosize_columns(ws)
    wb.save(output_path)

    # 項目別達成状況（items が未指定の場合は raw_df から生成）
    if items_df is None:
        items_df = make_sd_items(raw_df, is_simplified=is_simplified)
    items_records = df_to_records(items_df)

    # JSON output for dashboard
    dashboard_data = {
        "type": "symptom_disease",
        "stats": {
            "total_residents": len(summary_df),
            "avg_overall_rate": round(summary_df["全体達成率"].mean(), 3) if not summary_df.empty and pd.notna(summary_df["全体達成率"].mean()) else 0,
            "avg_approval_rate": round(summary_df["承認率"].mean(), 3) if "承認率" in summary_df.columns and not summary_df.empty and pd.notna(summary_df["承認率"].mean()) else 0,
            "total_alerts": len(alert_df)
        },
        "residents": df_to_records(summary_df),
        "alerts": df_to_records(alert_df),
        "items": items_records,
    }
    
    json_path = input_path.with_name(input_path.stem + "_visualized.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    return output_path, json_path


# ---------- evaluation ----------
BASE_COLS = ["研修医UMIN ID","研修医氏名","施設名","診療科名","研修開始日","研修終了日","主／並行","指導医UMIN ID","指導医氏名"]

def find_header_row(ws, search_rows: int = 10) -> Optional[int]:
    for r in range(1, min(ws.max_row, search_rows) + 1):
        row_vals = [cell_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "研修医UMIN ID" in row_vals and "研修医氏名" in row_vals:
            return r
    return None

EXCLUDED_SHEET_KEYWORDS = ["メディカルスタッフ", "患者・家族等", "患者家族"]

def read_all_eval_sheets(input_path: Path) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(input_path, data_only=True)
    eval_dfs = {}
    for ws in wb.worksheets:
        # 「メディカルスタッフ」「患者・家族等」など対象外シートはスキップ
        if any(kw in ws.title for kw in EXCLUDED_SHEET_KEYWORDS):
            continue
        header_row = find_header_row(ws)
        if not header_row:
            continue
        data = list(ws.values)
        header = [cell_text(v) for v in data[header_row - 1]]
        rows = data[header_row:]
        df = pd.DataFrame(rows, columns=header).dropna(how="all").copy()
        df.columns = [cell_text(c) for c in df.columns]
        if "研修医UMIN ID" in df.columns:
            df = df[df["研修医UMIN ID"].astype(str) != "研修医UMIN ID"].copy()

        # Rename customized column headers in the new format to match existing logic
        rename_map = {
            "研修施設": "施設名",
            "診療科": "診療科名",
            "評価者氏名": "指導医氏名",
            "評価者UMIN ID": "指導医UMIN ID"
        }
        df.rename(columns=rename_map, inplace=True)
        
        # Determine logical source name
        title = ws.title
        if "指導" in title:
            source_name = "指導医評価"
        elif "研修医" in title or "自己" in title:
            source_name = "研修医評価"
        else:
            source_name = title
            
        # Avoid overriding if same named logical sheet exists
        if source_name in eval_dfs:
            source_name = f"{source_name}_{title}"
            
        eval_dfs[source_name] = df.reset_index(drop=True)
        
    return eval_dfs

def get_score_columns(df: pd.DataFrame) -> List[str]:
    import re
    # A-1, B-2, C-3 のような上位項目のみ対象。B-1-1 等のサブ項目は除外。
    pattern = re.compile(r'^[ABC]-\d+\.')
    return [c for c in df.columns if pattern.match(c)]

def classify_competency(col: str) -> str:
    if col.startswith("A-"):
        return "A_プロフェッショナリズム"
    if col.startswith("B-"):
        return "B_資質・能力"
    if col.startswith("C-"):
        return "C_基本的診療業務"
    return "その他"

def to_long_eval(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    if df.empty:
        return df
    score_cols = get_score_columns(df)
    base_cols = [c for c in BASE_COLS if c in df.columns]
    long_df = df.melt(id_vars=base_cols, value_vars=score_cols, var_name="評価項目", value_name="評価点").copy()
    long_df["評価点_num"] = long_df["評価点"].apply(safe_float)
    long_df["評価群"] = long_df["評価項目"].apply(classify_competency)
    long_df["評価元"] = source_name
    return long_df

def make_eval_resident_summary(long_all: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (rid, rname), sub in long_all.groupby(["研修医UMIN ID","研修医氏名"], dropna=False):
        # 0点は評価対象外として除外し、有効スコアのみで平均を算出
        valid = sub[sub["評価点_num"].notna() & (sub["評価点_num"] > 0)]
        # 低評価件数は指導医評価のみを対象とする
        valid_supervisor = valid[valid["評価元"] == "指導医評価"] if "評価元" in valid.columns else pd.DataFrame()
        rows.append({
            "研修医UMIN ID": rid,
            "研修医氏名": rname,
            "評価件数": valid["評価点_num"].notna().sum(),
            "平均点_全体": round(valid["評価点_num"].mean(), 3) if valid["評価点_num"].notna().any() else None,
            "平均点_A": round(valid.loc[valid["評価群"]=="A_プロフェッショナリズム","評価点_num"].mean(), 3) if valid.loc[valid["評価群"]=="A_プロフェッショナリズム","評価点_num"].notna().any() else None,
            "平均点_B": round(valid.loc[valid["評価群"]=="B_資質・能力","評価点_num"].mean(), 3) if valid.loc[valid["評価群"]=="B_資質・能力","評価点_num"].notna().any() else None,
            "平均点_C": round(valid.loc[valid["評価群"]=="C_基本的診療業務","評価点_num"].mean(), 3) if valid.loc[valid["評価群"]=="C_基本的診療業務","評価点_num"].notna().any() else None,
            "最低点": round(valid["評価点_num"].min(), 3) if valid["評価点_num"].notna().any() else None,
            "2点未満件数": int((valid_supervisor["評価点_num"] < 2).sum()) if not valid_supervisor.empty else 0,
            "1点以下件数": int((valid_supervisor["評価点_num"] <= 1).sum()) if not valid_supervisor.empty else 0,
            "0点件数": int((sub["評価点_num"] == 0).sum()),  # 元データから0点件数は記録
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["平均点_全体","2点未満件数","0点件数"], ascending=[True, False, False]).reset_index(drop=True)
    return out

def make_eval_department_summary(long_all: pd.DataFrame) -> pd.DataFrame:
    if long_all.empty:
        return pd.DataFrame()
    # 0点は評価対象外として除外
    filtered = long_all[long_all["評価点_num"].notna() & (long_all["評価点_num"] > 0)]
    if filtered.empty:
        return pd.DataFrame()
    grp = filtered.groupby(["施設名","診療科名","評価元"], dropna=False)["評価点_num"]
    out = grp.agg(["count","mean","min","max"]).reset_index()
    out.columns = ["施設名","診療科名","評価元","評価件数","平均点","最低点","最高点"]
    out["平均点"] = out["平均点"].round(3)
    return out.sort_values(["平均点","評価件数"], ascending=[True, False]).reset_index(drop=True)

def make_eval_evaluator_summary(long_all: pd.DataFrame) -> pd.DataFrame:
    # 0点は評価対象外として除外
    filtered = long_all[long_all["評価点_num"].notna() & (long_all["評価点_num"] > 0)]
    if filtered.empty:
        return pd.DataFrame(columns=["指導医UMIN ID","指導医氏名","評価元","評価件数","平均点"])
    grp = filtered.groupby(["指導医UMIN ID","指導医氏名","評価元"], dropna=False)["評価点_num"]
    out = grp.agg(["count","mean"]).reset_index()
    out.columns = ["指導医UMIN ID","指導医氏名","評価元","評価件数","平均点"]
    out["平均点"] = out["平均点"].round(3)
    return out.sort_values(["平均点","評価件数"], ascending=[True, False]).reset_index(drop=True)

def make_eval_radar_source(long_all: pd.DataFrame) -> pd.DataFrame:
    # 0点は評価対象外として除外
    filtered = long_all[long_all["評価点_num"].notna() & (long_all["評価点_num"] > 0)]
    if filtered.empty:
        return pd.DataFrame(columns=["研修医UMIN ID","研修医氏名","評価元","評価群","平均点"])
    out = filtered.groupby(["研修医UMIN ID","研修医氏名","評価元","評価群"], dropna=False)["評価点_num"].mean().reset_index()
    out["平均点"] = out["評価点_num"].round(3)
    return out.drop(columns=["評価点_num"]).sort_values(["研修医氏名","評価元","評価群"]).reset_index(drop=True)

def make_rotation_status(long_all: pd.DataFrame) -> pd.DataFrame:
    """研修医ごと・ローテーションごとの入力状況を集計する。
    各行が1ローテーション = (研修医, 施設, 診療科, 研修期間) に対応し、
    研修医評価と指導医評価それぞれの入力有無・平均点・乖離状態を付与する。

    【注意】研修医評価シートには指導医名が入力されていない場合があるため、
    指導医UMIN ID / 指導医氏名 はマッチングキーから除外し、
    指導医名は指導医評価シート側のデータから取得する。
    """
    if long_all.empty:
        return pd.DataFrame()

    # ローテーションを一意に特定するキー（指導医情報は含めない）
    match_cols = [c for c in ["研修医UMIN ID", "研修医氏名", "施設名", "診療科名",
                               "研修開始日", "研修終了日"]
                  if c in long_all.columns]

    # 表示用の付加列（どちらかのシートに存在すれば採用）
    extra_cols = [c for c in ["主／並行"] if c in long_all.columns]

    # ローテーション × 評価元 ごとの平均点・件数（match_cols でグループ）
    agg = (long_all
           .groupby(match_cols + ["評価元"], dropna=False)["評価点_num"]
           .agg(avg="mean", cnt="count")
           .reset_index())
    agg["avg"] = agg["avg"].round(3)

    # 全ローテーション一覧（match_cols で重複除去。extra_cols は先着を採用）
    all_rotations = long_all[match_cols + extra_cols].drop_duplicates(subset=match_cols).copy()

    # 指導医名・指導医 ID は指導医評価シートから取得（研修医評価シートは空のことがあるため）
    supervisor_cols = [c for c in ["指導医UMIN ID", "指導医氏名"] if c in long_all.columns]
    if supervisor_cols:
        sup_info = (long_all[long_all["評価元"] == "指導医評価"]
                    .groupby(match_cols, dropna=False)[supervisor_cols]
                    .first()
                    .reset_index())
        # マージ前に型を統一（UMIN IDがfloat64になる場合があるためstr変換）
        for col in match_cols:
            all_rotations[col] = all_rotations[col].astype(str)
            sup_info[col] = sup_info[col].astype(str)
        all_rotations = all_rotations.merge(sup_info, on=match_cols, how="left")

    # 研修医評価 / 指導医評価 を横展開してマージ
    def _side(source_label: str, avg_col: str, cnt_col: str) -> pd.DataFrame:
        sub = agg[agg["評価元"] == source_label][match_cols + ["avg", "cnt"]].copy()
        for col in match_cols:
            sub[col] = sub[col].astype(str)
        return sub.rename(columns={"avg": avg_col, "cnt": cnt_col})

    # all_rotations の match_cols を str に揃える（sup_info マージ未実施の場合も考慮）
    for col in match_cols:
        all_rotations[col] = all_rotations[col].astype(str)

    merged = (all_rotations
              .merge(_side("研修医評価", "研修医評価_平均点", "研修医評価_件数"), on=match_cols, how="left")
              .merge(_side("指導医評価", "指導医評価_平均点", "指導医評価_件数"), on=match_cols, how="left"))

    # 入力有無フラグ
    merged["研修医評価あり"] = merged["研修医評価_件数"].fillna(0) > 0
    merged["指導医評価あり"] = merged["指導医評価_件数"].fillna(0) > 0

    def _status(row: pd.Series) -> str:
        r, s = row["研修医評価あり"], row["指導医評価あり"]
        if r and s:
            return "両方入力済"
        if r and not s:
            return "研修医のみ"
        if not r and s:
            return "指導医のみ"
        return "未入力"

    merged["入力状況"] = merged.apply(_status, axis=1)

    # 日付を文字列に変換（JSON シリアライズ対応）
    for col in ["研修開始日", "研修終了日"]:
        if col in merged.columns:
            merged[col] = pd.to_datetime(merged[col], errors="coerce").dt.strftime("%Y-%m-%d").where(
                pd.to_datetime(merged[col], errors="coerce").notna(), other=None)

    sort_cols = [c for c in ["研修医氏名", "研修開始日", "診療科名"] if c in merged.columns]
    if sort_cols:
        merged = merged.sort_values(sort_cols).reset_index(drop=True)

    return merged

def make_eval_item_scores(long_all: pd.DataFrame) -> pd.DataFrame:
    """評価項目ごと・評価元ごとの平均点を集計する。
    (研修医UMIN ID, 研修医氏名, 評価元, 評価群, 評価項目) でグループ化。
    0点は評価対象外として除外する。
    """
    if long_all.empty:
        return pd.DataFrame()
    # 0点は評価対象外として除外
    filtered = long_all[long_all["評価点_num"].notna() & (long_all["評価点_num"] > 0)]
    if filtered.empty:
        return pd.DataFrame()
    grp_cols = [c for c in ["研修医UMIN ID", "研修医氏名", "評価元", "評価群", "評価項目"]
                if c in filtered.columns]
    out = (filtered.groupby(grp_cols, dropna=False)["評価点_num"]
           .agg(平均点="mean", 件数="count")
           .reset_index())
    out["平均点"] = out["平均点"].round(3)
    return out.sort_values(grp_cols).reset_index(drop=True)

def make_eval_alerts(long_all: pd.DataFrame, low_threshold: float = 2.0) -> pd.DataFrame:
    # 0点は評価対象外のため除外し、1点以上2点未満のみアラート対象とする
    low_df = long_all[long_all["評価点_num"].notna() & (long_all["評価点_num"] > 0) & (long_all["評価点_num"] < low_threshold)].copy()
    if low_df.empty:
        return pd.DataFrame(columns=["評価元","研修医氏名","研修医UMIN ID","施設名","診療科名","指導医氏名","評価群","評価項目","評価点","状態"])
    low_df["状態"] = "低評価(<2)"
    # Avoid duplicate column '評価点' by dropping the original '評価点' text column before renaming
    out = low_df.drop(columns=["評価点"]).rename(columns={"評価点_num":"評価点"})[
        ["評価元","研修医氏名","研修医UMIN ID","施設名","診療科名","指導医氏名","評価群","評価項目","評価点","状態"]
    ]
    return out.sort_values(["評価点","研修医氏名","評価元","評価項目"], ascending=[True, True, True, True]).reset_index(drop=True)

def write_eval_output(input_path: Path, raw_resident, raw_supervisor, long_resident, long_supervisor,
                      resident_summary, department_summary, evaluator_summary,
                      radar_source, item_scores, rotation_status, alerts) -> Tuple[Path, Path]:
    output_path = input_path.with_name(input_path.stem + "_evaluation_visualized.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not raw_resident.empty: raw_resident.to_excel(writer, index=False, sheet_name="raw_研修医評価")
        if not raw_supervisor.empty: raw_supervisor.to_excel(writer, index=False, sheet_name="raw_指導医評価")
        if not long_resident.empty: long_resident.to_excel(writer, index=False, sheet_name="long_研修医評価")
        if not long_supervisor.empty: long_supervisor.to_excel(writer, index=False, sheet_name="long_指導医評価")
        resident_summary.to_excel(writer, index=False, sheet_name="resident_summary")
        department_summary.to_excel(writer, index=False, sheet_name="department_summary")
        evaluator_summary.to_excel(writer, index=False, sheet_name="evaluator_summary")
        radar_source.to_excel(writer, index=False, sheet_name="radar_source")
        if not item_scores.empty: item_scores.to_excel(writer, index=False, sheet_name="item_scores")
        if not rotation_status.empty: rotation_status.to_excel(writer, index=False, sheet_name="rotation_status")
        alerts.to_excel(writer, index=False, sheet_name="alerts")
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        style_header(ws, 1)
        autosize_columns(ws)
        apply_numeric_fill(ws)
    ws = wb.create_sheet("dashboard")
    ws["A1"] = "EPOC 評価ダッシュボード"; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "全体サマリー"; ws["A3"].font = BOLD
    # 低評価件数は指導医評価のみを対象とする
    alerts_supervisor = alerts[alerts["評価元"] == "指導医評価"] if not alerts.empty and "評価元" in alerts.columns else pd.DataFrame()
    stats = [
        ("研修医数", len(resident_summary)),
        ("研修医平均点(全体平均)", round(resident_summary["平均点_全体"].mean(), 3) if not resident_summary.empty else None),
        ("低評価件数(<2)", len(alerts_supervisor)),
        ("0点件数", int((alerts["状態"] == "0点").sum()) if not alerts.empty else 0),
    ]
    r = 4
    for k,v in stats:
        ws[f"A{r}"] = k; ws[f"B{r}"] = v; r += 1
    autosize_columns(ws)
    apply_numeric_fill(ws)
    wb.save(output_path)

    # JSON output for dashboard（低評価件数は指導医評価のみ）
    dashboard_data = {
        "type": "evaluation",
        "stats": {
            "total_residents": len(resident_summary),
            "avg_overall_score": round(resident_summary["平均点_全体"].mean(), 3) if not resident_summary.empty and pd.notna(resident_summary["平均点_全体"].mean()) else None,
            "total_low_evals": len(alerts_supervisor),
            "total_zero_evals": int((alerts["状態"] == "0点").sum()) if not alerts.empty else 0
        },
        "residents": df_to_records(resident_summary),
        "departments": df_to_records(department_summary),
        "evaluators": df_to_records(evaluator_summary),
        "radar": df_to_records(radar_source),
        "item_scores": df_to_records(item_scores),
        "rotations": df_to_records(rotation_status),
        "alerts": df_to_records(alerts)
    }

    json_path = input_path.with_name(input_path.stem + "_evaluation_visualized.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    return output_path, json_path

# ---------- case summary ----------
def run_case_summary(input_path: Path) -> Tuple[Path, Path]:
    try:
        df = pd.read_excel(input_path)
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みに失敗しました: {e}")

    # ヘッダーに名前やIDがあるか確認し、無い場合はエラー
    if "研修医氏名" not in df.columns:
        raise ValueError("ヘッダーに「研修医氏名」が見つかりませんでした。病歴要約等入力状況一覧ではない可能性があります。")

    # 氏名やID以外の列を症候/疾患アイテムとみなす
    exclude_cols = ["研修医氏名", "UMIN ID"]
    item_cols = [c for c in df.columns if c not in exclude_cols and not str(c).startswith("Unnamed")]

    # 未・済のカウント
    summary_rows = []
    item_stats = {col: {"済": 0, "未": 0, "空欄": 0} for col in item_cols}

    for idx, row in df.iterrows():
        resident_name = str(row.get("研修医氏名", f"Unknown_{idx}")).strip()
        if resident_name == "nan" or not resident_name:
            continue
            
        umin_id = str(row.get("UMIN ID", "")).strip()
        if umin_id == "nan":
            umin_id = ""

        completed = 0
        uncompleted = 0
        blank = 0
        
        for col in item_cols:
            val = str(row.get(col, "")).strip()
            if val == "済":
                completed += 1
                item_stats[col]["済"] += 1
            elif val == "未":
                uncompleted += 1
                item_stats[col]["未"] += 1
            else:
                blank += 1
                item_stats[col]["空欄"] += 1

        summary_rows.append({
            "研修医氏名": resident_name,
            "UMIN ID": umin_id,
            "入力済件数": completed,
            "未入力件数": uncompleted,
            "全体進捗率": round(completed / len(item_cols), 3) if item_cols else 0.0,
            "対象項目数": len(item_cols)
        })

    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(by=["入力済件数", "研修医氏名"], ascending=[False, True]).reset_index(drop=True)

    item_stats_list = []
    for col, counts in item_stats.items():
        total = counts["済"] + counts["未"]
        item_stats_list.append({
            "項目名": col,
            "済_件数": counts["済"],
            "未_件数": counts["未"],
            "入力率": round(counts["済"] / total, 3) if total > 0 else 0.0
        })
    item_stats_df = pd.DataFrame(item_stats_list).sort_values(by=["済_件数"], ascending=False).reset_index(drop=True)

    # 研修医×項目の済/未詳細（ダッシュボードの個別表示用）
    # 列インデックスで症候/疾患を判定（Excel AF列=32列目以降を疾患とする）
    col_to_idx = {c: i for i, c in enumerate(df.columns)}
    item_details = []
    for idx, row in df.iterrows():
        resident_name = str(row.get("研修医氏名", "")).strip()
        if resident_name == "nan" or not resident_name:
            continue
        for col in item_cols:
            val = str(row.get(col, "")).strip()
            status = "済" if val == "済" else "未"  # 空欄も未として扱う
            excel_col = col_to_idx.get(col, 0) + 1  # 1-based
            区分 = "疾患" if excel_col >= SYMPTOM_DISEASE_BOUNDARY_COL else "症候"
            item_details.append({
                "研修医氏名": resident_name,
                "項目名": col,
                "区分": 区分,
                "状態": status,
            })
    item_details_df = pd.DataFrame(item_details)

    # Output to Excel
    output_path = input_path.with_name(input_path.stem + "_case_summary_visualized.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="raw_data")
        summary_df.to_excel(writer, index=False, sheet_name="resident_summary")
        item_stats_df.to_excel(writer, index=False, sheet_name="item_summary")
        
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        style_header(ws, 1)
        autosize_columns(ws)
    apply_numeric_fill(wb["resident_summary"])
    
    ws = wb.create_sheet("dashboard")
    ws["A1"] = "EPOC 病歴要約等入力進捗ダッシュボード"; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "全体サマリー"; ws["A3"].font = BOLD
    stats = [
        ("研修医数", len(summary_df)),
        ("平均全体進捗率", round(summary_df["全体進捗率"].mean(), 3) if not summary_df.empty else 0),
        ("対象項目数", len(item_cols))
    ]
    r = 4
    for k,v in stats:
        ws[f"A{r}"] = k; ws[f"B{r}"] = v; r += 1
    autosize_columns(ws)
    wb.save(output_path)

    # Output to JSON
    dashboard_data = {
        "type": "case_summary",
        "stats": {
            "total_residents": len(summary_df),
            "avg_overall_progress": round(summary_df["全体進捗率"].mean(), 3) if not summary_df.empty and pd.notna(summary_df["全体進捗率"].mean()) else 0,
            "total_items": len(item_cols)
        },
        "residents": df_to_records(summary_df),
        "items": df_to_records(item_stats_df),
        "item_details": df_to_records(item_details_df),
    }

    json_path = input_path.with_name(input_path.stem + "_case_summary_visualized.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    return output_path, json_path



# ---------- simplified symptom_disease (from case input list) ----------
def run_simplified_symptom_disease(input_path: Path) -> Tuple[Path, Path]:
    try:
        df = pd.read_excel(input_path)
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みに失敗しました: {e}")

    if "研修医氏名" not in df.columns:
        raise ValueError("ヘッダーに「研修医氏名」が見つかりませんでした。")

    exclude_cols = ["研修医氏名", "UMIN ID"]
    item_cols = [c for c in df.columns if c not in exclude_cols and not str(c).startswith("Unnamed")]

    # 列インデックスを取得（Excel列AF=32以降を疾患とする）
    col_to_idx = {c: i for i, c in enumerate(df.columns)}

    summary_rows = []
    alert_rows = []
    heatmap_data = []

    for idx, row in df.iterrows():
        resident_name = str(row.get("研修医氏名", f"Unknown_{idx}")).strip()
        if resident_name == "nan" or not resident_name:
            continue
            
        umin_id = str(row.get("UMIN ID", "")).strip()
        if umin_id == "nan":
            umin_id = ""

        total_items = len(item_cols)
        experienced_count = 0
        symptom_count = sum(1 for c in item_cols if col_to_idx.get(c, 0) + 1 < SYMPTOM_DISEASE_BOUNDARY_COL)
        disease_count = total_items - symptom_count
        symptom_exp = 0
        disease_exp = 0
        
        for col in item_cols:
            val = row.get(col, 0)
            count = safe_int(val)
            # Excel列AF(32列目)以降を疾患、それ以前を症候とする
            excel_col_idx = col_to_idx.get(col, 0) + 1  # 1-based
            section = "疾患" if excel_col_idx >= SYMPTOM_DISEASE_BOUNDARY_COL else "症候"
            
            heatmap_data.append({
                "研修医氏名": resident_name,
                "研修医UMIN ID": umin_id,
                "区分": section,
                "項目名": col,
                "経験数": count,
                "承認数": count, # 簡易版に承認ステータスがないため経験数と同じとする
                "病歴要約提出": 0,
                "病歴要約確認": 0,
                "外科手術要約提出": 0,
                "外科手術要約確認": 0,
                "シート名": "簡易版一覧",
            })
            
            if count > 0:
                experienced_count += 1
                if section == "症候":
                    symptom_exp += 1
                else:
                    disease_exp += 1
            else:
                alert_rows.append({
                    "研修医氏名": resident_name,
                    "研修医UMIN ID": umin_id,
                    "区分": section,
                    "項目名": col,
                    "経験数": 0,
                    "承認数": 0,
                    "病歴要約提出": 0,
                    "病歴要約確認": 0,
                    "外科手術要約提出": 0,
                    "外科手術要約確認": 0,
                    "状態": "未経験"
                })

        summary_rows.append({
            "研修医氏名": resident_name,
            "研修医UMIN ID": umin_id,
            "総項目数": total_items,
            "経験済項目数": experienced_count,
            "承認済項目数": experienced_count, # 同上
            "症候達成率": round(symptom_exp / symptom_count, 3) if symptom_count > 0 else 0.0,
            "疾患達成率": round(disease_exp / disease_count, 3) if disease_count > 0 else 0.0,
            "全体達成率": round(experienced_count / total_items, 3) if total_items > 0 else 0.0,
            "承認率": round(experienced_count / total_items, 3) if total_items > 0 else 0.0,
            "病歴要約未提出件数": 0,
            "病歴要約未確認件数": 0,
            "外科手術要約未提出件数": 0,
            "外科手術要約未確認件数": 0,
            "要対応件数": total_items - experienced_count, # 未経験数 = 要対応とみなす
        })

    raw_df = pd.DataFrame(heatmap_data, columns=[
        "研修医氏名", "研修医UMIN ID", "区分", "項目名", "経験数", "承認数",
        "病歴要約提出", "病歴要約確認", "外科手術要約提出", "外科手術要約確認", "シート名"
    ])
    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(by=["要対応件数", "全体達成率"], ascending=[False, True]).reset_index(drop=True)

    alert_df = pd.DataFrame(alert_rows)
    if not alert_df.empty:
        alert_df = alert_df.sort_values(["研修医氏名", "項目名"])

    heat_symptom = pivot_sd_heatmap(raw_df, "症候")
    heat_disease = pivot_sd_heatmap(raw_df, "疾患") # empty DataFrame

    return write_sd_output(input_path, raw_df, summary_df, alert_df, heat_symptom, heat_disease, is_simplified=True)

# ---------- main ----------
def run_symptom_disease(input_path: Path) -> Tuple[Path, Path]:
    # 自動判定で簡易版だった場合は簡易版パースへルーティング
    filetype = detect_workbook_type(input_path)
    if filetype == "symptom_disease_simplified":
        return run_simplified_symptom_disease(input_path)
        
    raw_df = parse_symptom_disease_workbook(input_path)
    summary_df = make_sd_summary(raw_df)
    alert_df = make_sd_alerts(raw_df)
    heat_symptom = pivot_sd_heatmap(raw_df, "症候")
    heat_disease = pivot_sd_heatmap(raw_df, "疾患")
    return write_sd_output(input_path, raw_df, summary_df, alert_df, heat_symptom, heat_disease)

def run_evaluation(input_path: Path) -> Tuple[Path, Path]:
    eval_dfs = read_all_eval_sheets(input_path)
    if not eval_dfs:
        raise ValueError("評価ファイルとして有効なシート（研修医UMIN ID等の列を含む）を認識できませんでした。")
        
    long_dfs = []
    raw_resident = pd.DataFrame()
    raw_supervisor = pd.DataFrame()
    long_resident = pd.DataFrame()
    long_supervisor = pd.DataFrame()
    
    for source_name, df in eval_dfs.items():
        if df.empty:
            continue
        long_df = to_long_eval(df, source_name)
        long_dfs.append(long_df)
        
        if "指導" in source_name and raw_supervisor.empty:
            raw_supervisor = df
            long_supervisor = long_df
        elif raw_resident.empty:
            raw_resident = df
            long_resident = long_df

    if not long_dfs:
        raise ValueError("評価データが含まれていませんでした。")
        
    long_all = pd.concat(long_dfs, ignore_index=True)
    resident_summary = make_eval_resident_summary(long_all)
    department_summary = make_eval_department_summary(long_all)
    evaluator_summary = make_eval_evaluator_summary(long_all)
    radar_source = make_eval_radar_source(long_all)
    item_scores = make_eval_item_scores(long_all)
    rotation_status = make_rotation_status(long_all)
    alerts = make_eval_alerts(long_all)
    return write_eval_output(input_path, raw_resident, raw_supervisor, long_resident, long_supervisor,
                             resident_summary, department_summary, evaluator_summary,
                             radar_source, item_scores, rotation_status, alerts)

def main():
    parser = argparse.ArgumentParser(description="EPOC Auto Visualizer")
    parser.add_argument("input_file", help="Input Excel file path")
    parser.add_argument("--type", "-t", choices=["symptom_disease", "evaluation", "case_summary"], help="Explicitly specify the file type")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"ファイルが見つかりません: {input_path}")
        sys.exit(1)
        
    filetype = args.type
    if not filetype:
        filetype = detect_workbook_type(input_path)
        
    if filetype == "symptom_disease" or filetype == "symptom_disease_simplified":
        out, json_out = run_symptom_disease(input_path)
    elif filetype == "evaluation":
        out, json_out = run_evaluation(input_path)
    elif filetype == "case_summary":
        out, json_out = run_case_summary(input_path)
    else:
        raise ValueError("ファイル形式を判定できませんでした。明示的に --type (symptom_disease / evaluation / case_summary) を指定してください。")
    print(f"detected_type={filetype}")
    print(f"output={out}")
    print(f"json_output={json_out}")

if __name__ == "__main__":
    main()
