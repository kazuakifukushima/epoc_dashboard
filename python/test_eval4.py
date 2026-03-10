from openpyxl import load_workbook
wb = load_workbook("/Users/fukushimakazuaki/cursor/epoc_dashboard/P000005919_研修医評価票_example.xlsx", data_only=True)
ws = wb.worksheets[0]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(f"Row {i}:", [str(v)[:20] if v else None for v in row[:10]])
    if i > 5:
        break
